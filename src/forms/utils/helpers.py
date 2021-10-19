import os
from tempfile import NamedTemporaryFile
from typing import Union

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def normalize_answer(answer: Union[str, list]) -> str:
    """
    Line separate elements in answer if type is list
    """
    if isinstance(answer, list):
        answer = f'{os.linesep}'.join(answer)

    return answer


def stringify_bool(boolean: bool) -> str:
    """
    Return stringified version of boolean
    """
    return 'Да' if boolean else 'Нет'


def _separated_line_width(cell) -> int:
    """
    Calculate max width of cell for line separated str
    """
    width = max(map(len, cell.split()))

    return width


def adjust_width(ws: Worksheet, rows: list) -> None:
    """
    Set width for cells in worksheet
    """
    column_widths = []
    for row in rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                    if os.linesep not in str(cell):
                        column_widths[i] = len(str(cell))
                    else:
                        column_widths[i] = _separated_line_width(cell)
            else:
                column_widths += [len(str(cell))]
    for i, column_width in enumerate(column_widths):
        # Adding extra + 1 to width just in case
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 1


def _stylize_cells(ws: Worksheet) -> None:
    """
    Styling cells in worksheet
    """
    for idx, row in enumerate(ws.iter_rows()):
        for cell in row:
            if idx == 0:
                cell.font = Font(sz=12, b=True)
            cell.alignment = cell.alignment.copy(wrap_text=True)


def populate_ws(ws: Worksheet, rows: list) -> None:
    """
    Add data to worksheet
    """
    for row in rows:
        ws.append(row)

    _stylize_cells(ws)


def generate_report(filename, rows):
    workbook = Workbook()
    ws = workbook.active

    adjust_width(ws, rows)
    populate_ws(ws, rows)

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        stream = tmp.read()
        return HttpResponse(stream, headers={
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename={filename}.xlsx',
        })
