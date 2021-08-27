import datetime
import json
import os
from tempfile import NamedTemporaryFile
from typing import Union

import requests
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import HttpResponseNotFound, HttpResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from dashboard.forms import LeadForm
from dashboard.models import LessonUrl, Lead, Student, Course, Lesson, Promotion, Contact, Form
from dashboard.tasks import message_students_task, message_contacts_task, initiate_promo_task
from dashboard.utils.telegram import TelegramSender
from ffmpeg import get_resolution, get_duration

TELEGRAM_AGENT = 'TelegramBot (like TwitterBot)'
MESSAGE_URL = f'https://api.telegram.org/bot{os.getenv("BOT_TOKEN")}/sendMessage'
PHOTO_URL = f'https://api.telegram.org/bot{os.getenv("BOT_TOKEN")}/sendPhoto'


def normalize_answer(answer: Union[str, list]) -> str:
    """
    Line separate elements in answer if type is list
    """
    if isinstance(answer, list):
        answer = f'{os.linesep}'.join(answer)

    return answer


def stringify_bool(boolean: bool) -> str:
    """
    Return stringified version of boolean
    """
    return 'Да' if boolean else 'Нет'


def _separated_line_width(cell) -> int:
    """
    Calculate max width of cell for line separated str
    """
    width = max(map(len, cell.split()))

    return width


def adjust_width(ws: Worksheet, rows: list) -> None:
    """
    Set width for cells in worksheet
    """
    column_widths = []
    for row in rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    if os.linesep not in cell:
                        column_widths[i] = len(cell)
                    else:
                        column_widths[i] = _separated_line_width(cell)
            else:
                column_widths += [len(cell)]
    for i, column_width in enumerate(column_widths):
        # Adding extra + 1 to width just in case
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 1


def _stylize_cells(ws: Worksheet) -> None:
    """
    Styling cells in worksheet
    """
    for idx, row in enumerate(ws.iter_rows()):
        for cell in row:
            if idx == 0:
                cell.font = Font(sz=12, b=True)
            cell.alignment = cell.alignment.copy(wrap_text=True)


def populate_ws(ws: Worksheet, rows: list) -> None:
    """
    Add data to worksheet
    """
    for row in rows:
        ws.append(row)

    _stylize_cells(ws)


@login_required
def form_report(request, form_id: int):
    """
    Generate xlsx file for form
    """
    form = Form.objects.get(pk=form_id)
    questions = form.formquestion_set.all()
    answers = form.contactformanswers_set.all()
    workbook = Workbook()
    ws = workbook.active

    headers = ['Id', 'Студент', 'Дата прохождения', 'Зареган'] + [x.text for x in questions]
    collected_answers = []
    for answer in answers:
        date_passed = (answer.updated_at if answer.updated_at else answer.created_at).strftime('%m/%d/%Y, %H:%M')
        collected_answers.append(
            [str(answer.id), answer.contact.__str__(), date_passed, stringify_bool(answer.contact.is_registered)]
            + [normalize_answer(answer.data.get(str(x.id), '-')) for x in questions]
        )

    rows = [headers, *collected_answers]
    adjust_width(ws, rows)
    populate_ws(ws, rows)

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        stream = tmp.read()
        return HttpResponse(stream, headers={
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename={form.id}.xlsx',
        })


def watch_video(request, uuid):
    """
    Handle video request that comes from telegram bot
    """
    if request.META.get('HTTP_USER_AGENT') == TELEGRAM_AGENT:
        return HttpResponseNotFound()

    try:
        lesson_url = LessonUrl.objects.get(hash=uuid)
    except LessonUrl.DoesNotExist:
        msg = 'Ошибка 404. Данный урок не найден - возможно он был удален. ' \
              'Пожалуйста, зайдите в бот, откройте нужный курс и заново выберите требуемый урок.'
        return render(request, 'dashboard/error.html', {'error': msg})

    context = {'lesson': lesson_url.lesson}

    if lesson_url.hits >= 3:
        # Delete record from database once user opens the link
        lesson_url.delete()
        msg = 'Ошибка 403. Ссылка на урок устарела. Если вы хотите просмотреть этот урок, ' \
              'пожалуйста, зайдите в бот, откройте нужный курс и заново выберите требуемый урок.'
        return render(request, 'dashboard/error.html', {'error': msg})
    else:
        lesson_url.hits += 1
        lesson_url.save()
    return render(request, 'dashboard/watch_lesson.html', context)


def auth_and_watch(request, lesson_id):
    """
    Authorize user with telegram and provide access to video
    """

    # authentificating
    tg_id = request.GET.get('id')
    if tg_id and Student.objects.filter(tg_id=tg_id).exists():
        request.session['tg_id'] = tg_id
        request.session.set_expiry(60 * 60)
        return redirect('dashboard:auth_and_watch', lesson_id=lesson_id)

    # authentificated and all ok
    if request.session.get('tg_id'):
        student = Student.objects.get(tg_id=request.session.get('tg_id'))
        lesson = get_object_or_404(Lesson, pk=lesson_id)
        return render(request, 'dashboard/watch_lesson.html', {'lesson': lesson, 'student': student})
    else:
        return render(request, 'dashboard/tg_auth.html', {'id': lesson_id})


def signup(request):
    """
    Save lead from signup page
    """
    if request.POST:
        try:
            lead = Lead.objects.create(
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name'],
                language_type=request.POST['language_type'],
                phone=request.POST['phone'],
                city=request.POST['city'],
                chosen_field_id=request.POST['chosen_field'],
                application_type=Lead.ApplicationType.web,
                is_client=False
            )
            lead.refresh_from_db()
            return redirect(lead.invite_link)
        except IntegrityError:
            return HttpResponse('Invalid phone number. Number is already used')
    else:
        form = LeadForm()
        return render(request, 'dashboard/signup.html', {'form': form})


def message_to_students(request):
    """
    Handles message sending to students from Course in Admin panel
    """
    students = Student.objects.all()
    selected = getattr(request, request.method).getlist('_selected_action')
    referer = request.META['HTTP_REFERER']

    if selected:
        students = Student.objects.filter(pk__in=selected)

    course_id = getattr(request, request.method).get('course_id')
    lesson_id = getattr(request, request.method).get('lesson_id')

    if 'send' in request.POST:
        is_feedback = request.POST.get('is_feedback')

        config = {
            'is_feedback': is_feedback,
            'students': 'all' if not selected else selected,
            'message': request.POST['message'],
            'course_id': course_id,
            'lesson_id': lesson_id
        }

        message_students_task.delay(config)

        return HttpResponseRedirect(request.POST.get('referer'))

    return render(request, 'dashboard/send_intermediate.html', context={
        'entities': students,
        'course_id': course_id,
        'lesson_id': lesson_id,
        'referer': referer,
    })


def message_contacts(request):
    """
        Handles message sending to students from Course in Admin panel
        """
    contacts = Contact.objects.all()
    selected = getattr(request, request.method).getlist('_selected_action')
    referer = request.META['HTTP_REFERER']
    if selected:
        contacts = Contact.objects.filter(pk__in=selected)
    if 'send' in request.POST:
        selected = request.POST.getlist('_selected_action')

        is_feedback = request.POST.get('is_feedback')
        config = {
            'is_feedback': is_feedback,
            'contacts': 'all' if not selected else selected,
            'message': request.POST['message'],
        }

        message_contacts_task.delay(config)

        return HttpResponseRedirect(request.POST.get('referer'))

    return render(request, 'dashboard/send_intermediate.html', context={
        'entities': contacts,
        'referer': referer,
    })


def send_lesson(request, course_id, lesson_id):
    """
    Handles lesson sending to students from Course in Admin panel
    """
    course = Course.objects.get(pk=course_id)
    lesson = Lesson.objects.get(pk=lesson_id)
    students = course.student_set.all()

    kb = {'inline_keyboard': [[{'text': '', 'callback_data': f'data|lesson|{lesson.id}'}]]}

    url = MESSAGE_URL
    image = lesson.image.read() if lesson.image and not lesson.image_file_id else None

    for student in students:
        # todo studentlesson create record
        kb['inline_keyboard'][0][0]['text'] = 'Получить урок' if student.language_type == Student.LanguageType.ru \
            else 'Darsni tomosha qiling'
        data = {
            'chat_id': student.tg_id,
            'parse_mode': 'html',
            'text': render_to_string('dashboard/tg_lesson_info.html', {'lesson': lesson}),
            'reply_markup': json.dumps(kb)
        }

        files = None
        if lesson.image:
            url = PHOTO_URL
            data.pop('text')
            data['caption'] = render_to_string('dashboard/tg_lesson_info.html', {'lesson': lesson})
            if lesson.image_file_id:
                data['photo'] = lesson.image_file_id
            else:
                files = {'photo': image}

        resp = requests.post(url, data=data, files=files).json()
        if lesson.image and not lesson.image_file_id:
            lesson.image_file_id = resp['result']['photo'][-1]['file_id']

    lesson.date_sent = datetime.datetime.now()
    lesson.save()
    return HttpResponseRedirect(reverse('admin:dashboard_course_change', args=(course_id,)))


def send_promo(request, promo_id, lang):
    promotion = get_object_or_404(Promotion, pk=promo_id)
    message = render_to_string('dashboard/promo_text.html', {'promo': promotion})

    config = {
        'promo_id': promo_id,
        'message': message,
        'lang': lang
    }
    initiate_promo_task.delay(config)

    messages.add_message(request, messages.INFO, 'Отправлено всем студентам.')
    return HttpResponseRedirect(reverse('admin:dashboard_promotion_change', args=(promo_id,)))


def send_promo_myself(request, promo_id):
    promotion = get_object_or_404(Promotion, pk=promo_id)
    message = render_to_string('dashboard/promo_text.html', {'promo': promotion})
    image = promotion.image.path
    video = None
    thumb = None
    duration = None
    width = None
    height = None

    if promotion.video:
        image = None
        video = promotion.video.path
        thumb = promotion.image.path
        duration = get_duration(promotion.video.path)
        width, height = get_resolution(promotion.video.path)

    TelegramSender(int(os.getenv('CHAT_ID')), message, image, video,
                   duration, width, height, thumb).send()

    messages.add_message(request, messages.INFO, 'Отправлено в общий chat id.')
    return HttpResponseRedirect(reverse('admin:dashboard_promotion_change', args=(promo_id,)))
